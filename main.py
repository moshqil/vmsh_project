from openpyxl import Workbook, load_workbook

circles14_15_name = 'Kruzhki_2014-2015.xlsx'
trying15_name = 'Postupayuschie_v_mae_2015.xlsx'
finish15_name = 'Prinyatye_v_sentyabre_2015.xlsx'
circles15_16_name = 'Kruzhki_2015-2016.xlsx'
trying16_name = 'Postupayuschie_v_mae_2016.xlsx'
finish16_name = 'Prinyatye_v_sentyabre_2016.xlsx'
circles16_17_name = 'Kruzhki_2016-2017.xlsx'
trying17_name = 'Postupayuschie_2017.xlsx'
finish17_name = 'Postupivshie_v_2017.xlsx'


def open_excel(*names):
    return (load_workbook(name, read_only=True) for name in names)


def slice_values(slice_):
    return [cell[0].value for cell in slice_]


def split_names(names, i, j):
    return [(name.split()[i], name.split()[j]) for name in names]


def all_values(iterable):
    return (cell.value for cell in iterable)


def get_trying(ws, start, stop, i, j):
    people = set()
    for row in range(start, stop):
        people.add(tuple(map(str.capitalize, all_values((ws.cell(row=row, column=i), ws.cell(row=row, column=j))))))
    return people


def filter_good(ws, start, stop, i, k, good):
    people = set()
    for row in range(start, stop):
        if ws.cell(row=row, column=k).value == good:
            people.add(tuple(map(str.capitalize, ws.cell(row=row, column=i).value.split()[0:2])))
    return people


def stat(trying, finish, circles):
    a = len(trying)
    b = len(circles)
    c = len(trying & circles)
    d = len(finish & circles)
    e = len(finish)
    f = d / c
    return a, b, c, d, e, f


circles15, trying15, finish15 = open_excel(circles14_15_name, trying15_name, finish15_name)
finish15_7 = set(split_names(slice_values(finish15['Лист1']['B3':'B28']), 0, 1))
finish15_8 = set(split_names(slice_values(finish15['Лист1']['B31':'B55']), 0, 1))
finish15_9 = set(split_names(slice_values(finish15['Лист1']['B58':'B84']), 0, 1))
trying15_7 = get_trying(trying15['Поступающие в 7 класс'], 2, 249, 1, 2)
trying15_8 = get_trying(trying15['Поступающие в 8 класс'], 2, 253, 1, 2)
trying15_9 = get_trying(trying15['Поступающие в 9 класс'], 2, 198, 1, 2)
circles15_7 = get_trying(circles15['Математический 5-6'], 103, 171, 2, 3) | get_trying(circles15['Математический для 6 класса'], 1, 18, 1, 2)
circles15_8 = get_trying(circles15['Математический для 7 класса'], 1, 30, 1, 2)
circles15_9 = get_trying(circles15['Вечерний мат класс. 8 класс'], 1, 46, 1, 2)
print(stat(trying15_7, finish15_7, circles15_7))
print(stat(trying15_8, finish15_8, circles15_8))
print(stat(trying15_9, finish15_9, circles15_9))
circles16, trying16, finish16 = open_excel(circles15_16_name, trying16_name, finish16_name)
finish16_7 = set(split_names(slice_values(finish16['Лист1']['A1':'A29']), 0, 1))
finish16_8 = set(split_names(slice_values(finish16['Лист1']['A32': 'A60']), 0, 1))
finish16_9 = set(split_names(slice_values(finish16['Лист1']['A62': 'A84']), 0, 1))
trying16_7 = get_trying(trying16['В 7 класс'], 2, 266, 1, 2)
trying16_8 = get_trying(trying16['в 8 класс'], 2, 332, 1, 2)
trying16_9 = get_trying(trying16['в 9 класс'], 2, 105, 1, 2) | get_trying(trying16['в 9 класс'], 106, 204, 1, 2)
circles16_7 = filter_good(circles16['2015-2016'], 2, 462, 2, 4, 'Математический кружок для 5-6 классов') | filter_good(circles16['2015-2016'], 2, 462, 2, 4, 'Математический кружок. 6 класс. Продолжающий')
circles16_8 = filter_good(circles16['2015-2016'], 2, 462, 2, 4, 'Математический кружок для 7 класса')
circles16_9 = filter_good(circles16['2015-2016'], 2, 462, 2, 4, 'Вечерний математический класс. 8 класс') | filter_good(circles16['2015-2016'], 2, 462, 2, 4, 'Математический кружок для 8 класса')
print(stat(trying16_7, finish16_7, circles16_7))
print(stat(trying16_8, finish16_8, circles16_8))
print(stat(trying16_9, finish16_9, circles16_9))
circles17, trying17, finish17 = open_excel(circles16_17_name, trying17_name, finish17_name)
finish17_7 = set(split_names(slice_values(finish17['Лист1']['A1':'A27']), 0, 1))
finish17_8 = set(split_names(slice_values(finish17['Лист1']['A29': 'A55']), 0, 1))
finish17_9 = set(split_names(slice_values(finish17['Лист1']['A57': 'A83']), 0, 1))
trying17_7 = get_trying(trying17['Поступающие в 7 класс'], 2, 489, 2, 3)
trying17_8 = get_trying(trying17['Поступающие в 8 класс'], 2, 528, 1, 2)
trying17_9 = get_trying(trying17['Поступающие в 9 класс'], 2, 361, 1, 2)
circles17_7 = filter_good(circles17['Лист2'], 2, 748, 2, 4, 'Математический кружок для 5-6 классов')
circles17_8 = filter_good(circles17['Лист2'], 2, 748, 2, 4, 'Математический кружок для 7 класса (для начинающих и продолжающих)')
circles17_9 = filter_good(circles17['Лист2'], 2, 748, 2, 4, 'Вечерний математический класс. 8 класс')
print(stat(trying17_7, finish17_7, circles17_7))
print(stat(trying17_8, finish17_8, circles17_8))
print(stat(trying17_9, finish17_9, circles17_9))
